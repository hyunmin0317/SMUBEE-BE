from django.core.management import BaseCommand
from professors import models as professor_models
from openpyxl import load_workbook

duplicated_name = {
    "김영준": [
        ("HAAC0012", "yjnkim@smu.ac.kr"),
        ("HAEA0017", "yjkim@smu.ac.kr"),
        ("HAFL0002", "yjkim@smu.ac.kr"),
        ("HALR1221", "yjkim@smu.ac.kr"),
    ],
    "김은경": [
        ("HALF0122", "claire12@smu.ac.kr"),
        ("HALF5018", "claire12@smu.ac.kr"),
        ("HALF9024", "claire12@smu.ac.kr"),
        ("HALR1211", "claire12@smu.ac.kr"),
        ("HAAM9213", "claire12@smu.ac.kr"),
        ("HALF0601", "110083@smu.ac.kr"),
        ("HALF9061", "110083@smu.ac.kr"),
        ("HADA3301", "110083@smu.ac.kr"),
    ],
    "이성호": [
        ("HALR1214", "shlee@smu.ac.kr"),
        ("HABT0002", "shlee@smu.ac.kr"),
    ],
    "강상욱": [
        ("HACN0002", "swkang@smu.ac.kr"),
        ("HACN0030", "swkang@smu.ac.kr"),
        ("HAFX0009", "swkang@smu.ac.kr"),
        ("HALF9327", "swkang@smu.ac.kr"),
        ("HALR1213", "swkang@smu.ac.kr"),
        ("HALR1221", "sukang@smu.ac.kr"),
        ("HAEA0005", "sukang@smu.ac.kr"),
        ("HAEA0020", "sukang@smu.ac.kr"),
        ("HAEA9236", "sukang@smu.ac.kr"),
    ],
    "이상은": [
        ("HADA3341", "lse1213@smu.ac.kr"),
        ("HADA9226", "lse1213@smu.ac.kr"),
        ("HALR1225", "lse1213@smu.ac.kr"),
        ("HADA1102", "lse1213@smu.ac.kr"),
        ("HADA3321", "lse1213@smu.ac.kr"),
    ],
    "이형국": [
        ("HALF9390", "lhk@smu.ac.kr"),
        ("HALF9401", "lhk@smu.ac.kr"),
        ("HALF9402", "lhk@smu.ac.kr"),
    ],
}


class Command(BaseCommand):
    def handle(self, *args, **options):

        wb = load_workbook("2021-01-courses.xlsx")
        ws = wb.active

        for idx, row in enumerate(ws.iter_rows()):
            if idx == 0:
                continue

            pnames = row[3].value.split(",")
            print(pnames)
            emails = []
            for pname in pnames:
                professors = professor_models.Professor.objects.filter(name=pname)
                if len(professors) > 1:
                    course_code, _ = row[0].value.split("-")
                    for course in duplicated_name[pname]:
                        if course_code == course[0]:
                            emails.append(course[1])
                            break
                elif len(professors) == 1 and professors.first().email is not None:
                    emails.append(professors.first().email)
                else:
                    emails.append("NOTFOUND")
            ws.cell(idx + 1, 14, ",".join(emails))

        wb.save("preprocessed_courses_information.xlsx")
        wb.close()

        self.stdout.write(
            self.style.SUCCESS(f"Added professors' email to the each courses!")
        )
