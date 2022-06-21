from django.core.management import BaseCommand
from courses import models as course_models
from professors import models as professor_models
from openpyxl import load_workbook


class Command(BaseCommand):
    def handle(self, *args, **options):

        wb = load_workbook("preprocessed_courses_information.xlsx")
        ws = wb.active

        for idx, row in enumerate(ws.iter_rows()):
            if idx == 0:
                continue

            course_code, class_number = row[0].value.split("-")
            title = row[1].value

            course = course_models.Course.objects.create(
                title=title,
                course_code=course_code,
                class_number=class_number,
            )

            pfs = row[3].value.split(",")
            emails = row[13].value.split(",")
            for pf, email in zip(pfs, emails):
                if email != "NOTFOUND":
                    professor = professor_models.Professor.objects.get(
                        name=pf, email=email
                    )
                else:
                    try:
                        professor = professor_models.Professor.objects.get(name=pf)
                    except professor_models.Professor.DoesNotExist:
                        professor = professor_models.Professor.objects.create(name=pf)
                course.professors.add(professor)

            course.save()

        self.stdout.write(
            self.style.SUCCESS(f"Created Connection b/w professors and courses!")
        )
